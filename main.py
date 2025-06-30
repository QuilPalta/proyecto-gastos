import imaplib
import email
import re
import time
import requests
import os
import threading
from flask import Flask, request
from openpyxl import Workbook, load_workbook
from datetime import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# --- Configuración de entorno ---
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")
FOLDER_ID = os.getenv("GDRIVE_FOLDER_ID")

# Guardar credenciales.json desde variable base64
import base64
with open("credentials.json", "wb") as f:
    f.write(base64.b64decode(os.getenv("GDRIVE_CREDENTIALS_B64")))

# --- Excel ---
def guardar_en_excel(monto, comercio, fecha):
    filename = "gastos.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cargos"
        ws.append(["Fecha", "Comercio", "Monto", "Comentario"])
    else:
        wb = load_workbook(filename)
        ws = wb["Cargos"]

    fecha_formato = datetime.strptime(fecha, "%d/%m/%Y %H:%M").date()
    ws.append([fecha_formato, comercio, int(monto.replace(".", "")), "pendiente"])
    wb.save(filename)

# --- Drive ---
def subir_a_drive(nombre_archivo, archivo_local):
    creds = service_account.Credentials.from_service_account_file(
        "credentials.json", scopes=["https://www.googleapis.com/auth/drive"])
    service = build('drive', 'v3', credentials=creds)

    query = f"name='{nombre_archivo}' and trashed=false"
    if FOLDER_ID:
        query += f" and '{FOLDER_ID}' in parents"
    results = service.files().list(q=query, spaces='drive',
                                   fields="files(id, name)").execute()
    items = results.get('files', [])

    media = MediaFileUpload(archivo_local, resumable=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if items:
        service.files().update(fileId=items[0]['id'], media_body=media).execute()
    else:
        file_metadata = {'name': nombre_archivo}
        if FOLDER_ID:
            file_metadata['parents'] = [FOLDER_ID]
        service.files().create(body=file_metadata, media_body=media, fields='id').execute()

def obtener_link_archivo(nombre_archivo):
    creds = service_account.Credentials.from_service_account_file(
        "credentials.json", scopes=["https://www.googleapis.com/auth/drive"])
    service = build('drive', 'v3', credentials=creds)

    query = f"name='{nombre_archivo}' and trashed=false"
    if FOLDER_ID:
        query += f" and '{FOLDER_ID}' in parents"
    results = service.files().list(q=query, fields="files(id)").execute()
    items = results.get('files', [])
    if items:
        return f"https://drive.google.com/file/d/{items[0]['id']}/view?usp=drivesdk"
    return None

# --- Bot ---
app = Flask(__name__)

@app.route(f"/bot{TELEGRAM_TOKEN}", methods=["POST"])
def recibir_mensaje():
    data = request.json
    mensaje = data["message"]
    texto = mensaje.get("text", "")
    chat_id = mensaje["chat"]["id"]

    if texto.strip().lower() == "/excel":
        link = obtener_link_archivo("gastos.xlsx")
        mensaje = f"🧾 Aquí tienes el Excel con los gastos:
{link}" if link else "No se encontró el archivo."
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage", data={
            "chat_id": chat_id,
            "text": mensaje
        })
    return {"ok": True}

# --- Revisión de correos ---
def obtener_emails():
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    imap.login(EMAIL_USER, EMAIL_PASS)
    imap.select("inbox")
    status, messages = imap.search(None, '(FROM "enviodigital@bancochile.cl" SUBJECT "Cargo en Cuenta")')
    correos = []
    for num in messages[0].split()[-5:]:
        status, data = imap.fetch(num, "(RFC822)")
        msg = email.message_from_bytes(data[0][1])
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/html":
                    body = part.get_payload(decode=True).decode()
                    correos.append((num, body))
    imap.logout()
    return correos

def extraer_datos(html):
    match = re.search(r"una compra por \$([\d\.]+).*?en (.*?) el (\d{2}/\d{2}/\d{4} \d{2}:\d{2})", html)
    return match.groups() if match else (None, None, None)

def enviar_telegram(monto, comercio, fecha):
    mensaje = f"💳 Nuevo cargo:
💰 ${monto}
🏬 {comercio}
📅 {fecha}"
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    requests.post(url, data={"chat_id": TELEGRAM_CHAT_ID, "text": mensaje})

def loop_emails():
    enviados = set()
    while True:
        emails = obtener_emails()
        for id_correo, html in emails:
            if id_correo in enviados:
                continue
            monto, comercio, fecha = extraer_datos(html)
            if monto:
                enviar_telegram(monto, comercio, fecha)
                guardar_en_excel(monto, comercio, fecha)
                subir_a_drive("gastos.xlsx", "gastos.xlsx")
                enviados.add(id_correo)
        time.sleep(300)

threading.Thread(target=loop_emails, daemon=True).start()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)